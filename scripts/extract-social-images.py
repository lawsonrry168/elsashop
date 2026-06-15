"""Extract embedded images from Kang Zi Jian Social Content.xlsx."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Materials" / "2025 2026 康姿健 Social Content.xlsx"
OUT_DIR = ROOT / "public" / "images" / "social"
MANIFEST = ROOT / "src" / "data" / "social-images.json"

TOPIC_COL = 2
BODY_COL = 3
VISUAL_COL = 4
HASHTAG_COL = 5


def slugify(text: str, max_len: int = 80) -> str:
    text = re.sub(r"[^\w\u4e00-\u9fff\-]+", "-", text.strip())
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:max_len] or "image"


def main() -> int:
    if not XLSX.exists():
        print(f"Missing workbook: {XLSX}", file=sys.stderr)
        return 1

    if OUT_DIR.exists():
        for path in OUT_DIR.glob("*"):
            if path.is_file():
                path.unlink()
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(XLSX, data_only=True)
    items: list[dict] = []
    used_ids: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        images_by_row: dict[int, object] = {}
        for image in getattr(ws, "_images", []):
            row = image.anchor._from.row + 1  # noqa: SLF001
            images_by_row[row] = image

        for row_idx in range(2, ws.max_row + 1):
            topic = ws.cell(row=row_idx, column=TOPIC_COL).value
            body = ws.cell(row=row_idx, column=BODY_COL).value
            visual = ws.cell(row=row_idx, column=VISUAL_COL).value
            hashtag = ws.cell(row=row_idx, column=HASHTAG_COL).value

            if not topic and row_idx not in images_by_row:
                continue

            image = images_by_row.get(row_idx)
            if not image:
                continue

            topic_str = str(topic or "").strip()
            base_id = f"{sheet_name}-{slugify(topic_str)}"
            count = used_ids.get(base_id, 0)
            used_ids[base_id] = count + 1
            image_id = base_id if count == 0 else f"{base_id}-{count + 1}"

            ext = image.format.lower() if image.format else "png"
            filename = f"{image_id}.{ext}"
            filepath = OUT_DIR / filename
            filepath.write_bytes(image._data())  # noqa: SLF001

            items.append(
                {
                    "id": image_id,
                    "sheet": sheet_name,
                    "row": row_idx,
                    "file": f"/images/social/{filename}",
                    "topic": topic_str or None,
                    "body": str(body).strip() if body else None,
                    "visual": str(visual).strip() if visual else None,
                    "hashtag": str(hashtag).strip() if hashtag else None,
                }
            )

    manifest = {
        "source": "Materials/2025 2026 康姿健 Social Content.xlsx",
        "count": len(items),
        "items": items,
    }
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    logo_src = ROOT / "Materials" / "logo.jpg"
    logo_dst = ROOT / "public" / "brand" / "logo.jpg"
    if logo_src.exists():
        logo_dst.parent.mkdir(parents=True, exist_ok=True)
        logo_dst.write_bytes(logo_src.read_bytes())

    logo_script = ROOT / "scripts" / "remove-logo-background.py"
    if logo_script.exists():
        import subprocess
        import sys

        subprocess.run([sys.executable, str(logo_script)], check=False)

    print(f"images: {len(items)}")
    print(f"manifest: {MANIFEST}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
